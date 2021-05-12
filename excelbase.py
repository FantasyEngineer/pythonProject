from openpyxl import *


class ExcelManager(object):
    # def __init__(self):
    # self.file = file
    # self.wb = load_workbook(self.file)
    # sheets = self.wb.get_sheet_names()
    # self.sheet = sheets[0]
    # self.ws = self.wb[self.sheet]

    @staticmethod
    def getWorkBook(file):
        wb = load_workbook(file)
        return wb

    # 根据file以及sheet那么获取需要操作的sheet工作表
    @staticmethod
    def getSheet(file, sheetName):
        wb = ExcelManager.getWorkBook(file)
        iscon: bool = wb.sheetnames.__contains__(sheetName)
        if iscon:
            return wb[sheetName]
        else:
            return wb.create_sheet(sheetName)

    # 根据传入的sheet来 获取表格的总行数和总列数
    @staticmethod
    def get_row_clo_num(sheet):
        rows = sheet.max_row
        columns = sheet.max_column
        return rows, columns

    # 获取某个单元格的值
    def get_cell_value(self, row, column):
        cell_value = self.ws.cell(row=row, column=column).value
        return cell_value

    # 获取某列的所有值
    def get_col_value(self, column):
        rows = self.ws.max_row
        column_data = []
        for i in range(1, rows + 1):
            cell_value = self.ws.cell(row=i, column=column).value
            column_data.append(cell_value)
        return column_data

    # 获取某行所有值
    def get_row_value(self, row):
        columns = self.ws.max_column
        row_data = []
        for i in range(1, columns + 1):
            cell_value = self.ws.cell(row=row, column=i).value
            row_data.append(cell_value)
        return row_data

    # 设置某个单元格的值
    def set_cell_value(self, row, colunm, cellvalue):
        try:
            self.ws.cell(row=row, column=colunm).value = cellvalue
            self.wb.save(self.file)
        except:
            self.ws.cell(row=row, column=colunm).value = "writefail"
            self.wb.save(self.file)
