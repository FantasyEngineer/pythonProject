import tkinter

from openpyxl import *
from openpyxl.worksheet.worksheet import Worksheet
import time  # 引入time模块
import calendar  # 引入日历
import chinese_calendar
from chinese_calendar import is_workday
from datetime import datetime
from tkinter import *
from tkinter.filedialog import askopenfilename
from tkinter.messagebox import showinfo

# 全局化wb
from pip._internal.utils import subprocess

wb: Workbook
sheet: Worksheet


# 点击查看日历
def lookCal():
    try:
        # 获取到日历
        year = int(yearEntry.get())
        month = int(MonthEntry.get())
        cal = calendar.month(year, month)
        showinfo(message=cal.__str__())
    except Exception:
        showinfo(message="年月的输入有错误")


# 选择文件
def fileopen():
    file_sql = askopenfilename()
    if file_sql:
        filePath.set(file_sql)
        path = filePath.get()
        if path.__eq__('') or path is None:
            tkinter.messagebox.showwarning(message='表格的路径不能为空！')
            return
        try:
            # 加载表格,并且赋值道全局
            global wb
            wb = load_workbook(path)
            var2.set(wb.sheetnames)
            # 将sheet挂上
            frameSheet.pack(padx=10, pady=10, side=TOP, fill=X)
            # 将查找挂上
            runBtn1.pack(side=LEFT)
        except Exception:
            tkinter.messagebox.showwarning(message='宝，你好像选择的不是excel文件吧？ 要不再确认一下？')


secApartmentIndex: int = 0
bianzhiIndex: int = 0
enterTimeIndex: int = 0
outTimeIndex: int = 0


# 查找文字所在的列
def search():
    global sheet, secApartmentIndex, bianzhiIndex, enterTimeIndex, outTimeIndex
    # 获取用户选中的sheet
    sheetName: str = lb.get(lb.curselection())  # 获取当前选中的文本
    if sheetName.__eq__(''):
        tkinter.messagebox.showwarning(message='宝贝，你最少要选择一个sheet，我才能工作啊！')

    isApartMent: bool = 1
    isbianzhi: bool = 1
    isenter: bool = 1
    sheet = wb[sheetName]
    for i in range(1, sheet.max_row):
        for y in range(1, sheet.max_column):
            cell: str = sheet.cell(row=i, column=y).value
            if cell is None:
                continue
            if cell.__eq__("二级部门") and isApartMent:
                secApartmentIndex = y - 1
                isApartMent = 0
            elif cell.__eq__("编制") and isbianzhi:
                bianzhiIndex = y - 1
                isbianzhi = 0
            elif cell.__eq__("入职时间") and isenter:
                enterTimeIndex = y - 1
                isenter = 0
    # 展示数据
    outTimeIndex = enterTimeIndex + 1
    indexLabel = Label(window, text=f"二级部门{secApartmentIndex},编制{bianzhiIndex},入职时间{enterTimeIndex},离职时间{outTimeIndex}")
    indexLabel.pack(side=LEFT)

    # 将运行挂上
    runBtn.pack(side=BOTTOM, pady=20)


def run():
    print(f"二级部门{secApartmentIndex},编制{bianzhiIndex},入职时间{enterTimeIndex},离职时间{outTimeIndex}")
    try:
        print('---------------------时间计算----------------------------')
        # 获取到日历
        year = int(yearEntry.get())
        month = int(MonthEntry.get())

        cal = calendar.month(year, month)
        print(f"{year}年{month}月日历:")
        print(cal)
        # 计算当前月有多少天，用以获取最后一天的时间
        monthLen = calendar.monthlen(year, month)
        # 所以能计算出当前月的第一天，以及最后一天
        monthStartData = f"{year}/{month}/1"
        monthEndData = f"{year}/{month}/{monthLen}"
        # 将时间格式化成date格式
        monthStartData = datetime.strptime(monthStartData, '%Y/%m/%d')
        monthEndData = datetime.strptime(monthEndData, '%Y/%m/%d')

        # 获取所有的上班时间
        workDayAll = []
        # 获取当月的所有的上班日
        for i in range(1, monthLen + 1):
            day = f"{year}/{month}/{i}"
            day = datetime.strptime(day, '%Y/%m/%d')
            isWorkDay = chinese_calendar.is_workday(day)
            if isWorkDay:
                workDayAll.append(day)

        print(workDayAll)

        # 每周共有多少工作日
        workDays = len(workDayAll)
        # 每周的每人的工作总小时数，总工作日*8
        totalHourOnePersonMonth = workDays * 8

        print(f'当月开始时间：{monthStartData},当月结束时间：{monthEndData},当月需要工作天数：{workDays}')

        # 获取整行数据
        data_all = sheet.rows
        # 正编的小时数
        zhengbianTotalHour: int = 0
        # 外包的小时数
        waobaoTotalHour: int = 0
        # 符合要求的所有的人数
        totalPersonNum: int = 0

        # 元数组
        data_row = list(data_all)
        for row in data_row:
            # 获取二级部门，下标是1，编制在4，入职时间12，离职时间是13
            secDepartment: str = row[secApartmentIndex].value;
            # 姓名
            name: str = row[2].value
            # 编制
            bianzhi: str = row[bianzhiIndex].value
            # 入职时间
            enterTime: datetime = row[enterTimeIndex].value
            # 离职时间
            outTime: datetime = row[outTimeIndex].value

            # print(secDepartment)
            # 当不是这几个部门的时间，再开始计算这个人的当月的小时数
            if not '人事'.__eq__(secDepartment) and not '业务'.__eq__(secDepartment) and not '二级部门'.__eq__(
                    secDepartment) and secDepartment is not None:
                # 计算一下人数，也可以看到当前执行的位置
                totalPersonNum += 1
                # 输出当前行的所有数据
                # print(secDepartment, name, bianzhi, enterTime, outTime)
                if bianzhi.__eq__("正编"):  # 正编
                    # 当是在职状态下的时候，直接将在职改成下个月当天的时间。便于处理
                    if outTime.__str__().strip().__eq__("在职"):
                        outTime = monthEndData

                    if outTime >= monthEndData and enterTime < monthStartData:
                        # print(name, "离职时间是大于当月的，或者是月末最后一天，但是全勤")
                        zhengbianTotalHour += totalHourOnePersonMonth
                    else:
                        # 当月离职的。
                        if enterTime < monthStartData <= outTime < monthEndData:
                            print(name, "当月离职的", outTime)
                            for i in range(len(workDayAll)):
                                if outTime >= workDayAll[i]:
                                    # print(name, outTime, workDayAll[i])
                                    zhengbianTotalHour += 8
                        # 说明是当月就职的,并且当月不离职
                        elif monthStartData <= enterTime <= monthEndData < outTime:
                            # print(name, f"{month}月入职")
                            for i in range(len(workDayAll) - 1, 0, -1):
                                if enterTime <= workDayAll[i]:
                                    # print(name, enterTime, workDayAll[i])
                                    zhengbianTotalHour += 8
                        # 当月就职，当月离职
                        elif monthStartData <= enterTime <= monthEndData and outTime <= monthEndData:
                            # print(name, f"{month}月入职，当月离职了")
                            for i in range(len(workDayAll) - 1, 0, -1):
                                if enterTime <= workDayAll[i] <= outTime:
                                    # print(name, enterTime, workDayAll[i])
                                    zhengbianTotalHour += 8
                else:  # 外包
                    # 当是在职状态下的时候，直接将在职改成下个月当天的时间。便于处理
                    if outTime.__str__().strip().__eq__("在职"):
                        outTime = monthEndData

                    if outTime >= monthEndData and enterTime < monthStartData:
                        # print(name, "离职时间是大于当月的，或者是月末最后一天，但是全勤")
                        waobaoTotalHour += totalHourOnePersonMonth
                    else:
                        # 当月离职的。
                        if enterTime < monthStartData <= outTime < monthEndData:
                            # print(name, "当月离职的", outTime)
                            for i in range(len(workDayAll)):
                                if outTime >= workDayAll[i]:
                                    # print(name, outTime, workDayAll[i])
                                    waobaoTotalHour += 8
                        # 说明是当月就职的,并且当月不离职
                        elif monthStartData <= enterTime <= monthEndData < outTime:
                            # print(name, f"{month}月入职")
                            for i in range(len(workDayAll) - 1, 0, -1):
                                if enterTime <= workDayAll[i]:
                                    # print(name, enterTime, workDayAll[i])
                                    waobaoTotalHour += 8
                        # 当月就职，当月离职
                        elif monthStartData <= enterTime <= monthEndData and outTime <= monthEndData:
                            # print(name, f"{month}月入职，当月离职了")
                            for i in range(len(workDayAll) - 1, 0, -1):
                                if enterTime <= workDayAll[i] <= outTime:
                                    # print(name, enterTime, workDayAll[i])
                                    waobaoTotalHour += 8

        showinfo(message=f'去除人事以及业务总人数是{totalPersonNum}\n'
                         f'{month}月满勤共有{workDays}天上班时间，满勤总小时数是：{totalHourOnePersonMonth}\n'
                         f'{month}月正编工时数{zhengbianTotalHour}\n'
                         f'{month}月外包工时数{waobaoTotalHour}\n'
                         f'{month}月总工时数{waobaoTotalHour + zhengbianTotalHour}')
    except Exception:
        showinfo(message="运行出现了一点点问题，这里需要与0602sheet的格式保持一致")


# 创建窗体
window = Tk()
# 创建标题
window.title('复兴总工时处理程序')
window.configure(bg='white')
# 创建大小，以及位移
window.geometry('800x400+300+200')

# ------------------------------创建处理总工时的年月日容器------------------------------

frameTime = Frame(window)
frameTime.configure(bg='white')

tip = Label(frameTime, text='1.填写需要整理的日期：')
tip.pack(side=LEFT)

yearEntry = Entry(frameTime, justify=CENTER)
yearEntry.pack(side=LEFT)
yearEntry.insert(0, datetime.now().year)

yearlaber = Label(frameTime, text='年', bg='white')
yearlaber.pack(side=LEFT)

MonthEntry = Entry(frameTime, justify=CENTER)
MonthEntry.pack(side=LEFT)
MonthEntry.insert(0, datetime.now().month - 1)

monthlaber = Label(frameTime, text='月', bg='white')
monthlaber.pack(side=LEFT)

lookCalBtn = Button(frameTime, text="查看日历", width='30', command=lookCal)
lookCalBtn.pack(side=LEFT)

frameTime.pack(padx=10, pady=10, fill=X)

# ------------------------------创建选择文件的容器------------------------------
frameSelectFile = Frame(window)
frameSelectFile.pack(padx=10, pady=10, fill=X)
frameSelectFile.configure(bg='white')
# 创建button
btn = Button(frameSelectFile, text='2.点击选择操作的文件', command=fileopen)
btn.pack(side=LEFT)
# 创建展示文件夹路径的输入框
filePath = StringVar()
filePathEntry = Entry(frameSelectFile, width=100, textvariable=filePath)
filePathEntry.pack(padx=8, side=RIGHT, fill=X)

# ------------------------------创建展示sheets的listbox------------------------------
frameSheet = Frame(window)
sheetLabel = Label(frameSheet, text='3.请选择需要操作的sheet：')
sheetLabel.pack(side=LEFT)
# 创建sheetNames所在的Listbox
var2 = StringVar()
lb = Listbox(frameSheet, listvariable=var2)  # 将var2的值赋给Listbox
lb.pack(side=TOP, fill=X)

# ------------------------------创建运行按钮------------------------------
runBtn = Button(window, text="运行", width='30', bg='red', command=run)
runBtn1 = Button(window, text="查找所在字符所在的列数", width='30', command=search)

# 窗体循环
window.mainloop()
