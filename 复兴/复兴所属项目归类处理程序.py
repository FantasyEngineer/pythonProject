import tkinter
import _thread

from openpyxl import *
from openpyxl.worksheet.worksheet import *
import time  # 引入time模块
import calendar  # 引入日历
import chinese_calendar
from chinese_calendar import is_workday
from datetime import datetime
from tkinter import *
from tkinter.filedialog import askopenfilename
from tkinter.messagebox import showinfo

# 全局化wb
from pip._internal.utils import subprocess

平台优化改版: str = "技术-Thomas Cook平台" \
              "UED设计管理" \
              "产品设计管理" \
              "运营设计需求申请" \
              "2020线上问题跟踪" \
              "2021线上问题跟踪" \
              "产品设计管理 - 已迁移至智能科创需求管理" \
              "运营设计需求申请"
智能科创优化: str = "公共项目研发" \
              "科创需求项目集" \
              "技术-复游城" \
              "技术-集团IT管理" \
              "技术-亚特协同项目" \
              "科创管理" \
              "智能科创开发测试缺陷跟踪 " \
              "技术-数据中台项目" \
              "FTG-数据中台需求" \
              "技术-技术中台项目" \
              "技术-质量管理" \
              "请假与例会" \
              "请假" \
              "MBP项目" \
              "10-TCP平台" \
              "CTO" \
              "许愿池"
TC_UK: str = "技术-TCT UK协同与支持 TC APP Project"
酒店及旅行社相关: str = "酷怡 技术-酒店整体解决方案 技术-FOTOUR平台 技术-FOTEL解决方案 酷怡业务需求 FOTEL Lijiang 新业务研发 O2O项目（智能管家&码上游）"


# 选择文件
def fileopen():
    file_sql = askopenfilename()
    if file_sql:
        filePath.set(file_sql)

        filePathList = file_sql.split('.')
        print(filePathList)
        name: str = filePathList[0]
        suffix: str = filePathList[1]

        filePathSave.set(name + "_modify." + suffix)


def run():
    loading
    _thread.start_new_thread(run1, tuple())


def run1():
    if filePath.get() is None:
        showinfo("请确认文件路径是否正确")
        return
    # 读取表格
    wb = load_workbook(filePath.get())
    # 取到sheet
    sheet: Worksheet = wb['数据源']

    # 获取最大列数，创建比最大列还要打的一列
    columnNew = sheet.max_column + 1
    sheet.insert_cols(columnNew)
    isProject = 1
    projectIndex = 0
    # 查找所属项目所在的列
    for i in range(1, sheet.max_row):
        for y in range(1, sheet.max_column):
            cell: str = sheet.cell(row=i, column=y).value
            if cell is None:
                continue
            if cell.__eq__("所属项目") and isProject:
                projectIndex = y
                isProject = 0

    # 获取整行的数据
    for i in range(1, sheet.max_row + 1):
        cell: Cell = sheet.cell(row=i, column=projectIndex)
        value: str = cell.value
        # print(cell, value)
        if value is None:
            sheet.cell(row=i, column=columnNew).value = ""
        elif 平台优化改版.__contains__(value):
            sheet.cell(row=i, column=columnNew).value = "平台优化改版"
        elif 智能科创优化.__contains__(value):
            sheet.cell(row=i, column=columnNew).value = "智能科创优化"
        elif TC_UK.__contains__(value):
            sheet.cell(row=i, column=columnNew).value = "TC_UK"
        elif 酒店及旅行社相关.__contains__(value):
            sheet.cell(row=i, column=columnNew).value = "酒店及旅行社相关"

    wb.save(filePathSave.get())
    showinfo(message="处理成功")


def formatForm(form, width, heigth):
    """设置居中显示"""
    # 得到屏幕宽度
    win_width = form.winfo_screenwidth()
    # 得到屏幕高度
    win_higth = form.winfo_screenheight()

    # 计算偏移量
    width_adjust = (win_width - width) / 2
    higth_adjust = (win_higth - heigth) / 2

    form.geometry("%dx%d+%d+%d" % (width, heigth, width_adjust, higth_adjust))


# 创建窗体
window = Tk()
# 创建标题
window.title('复兴所属项目归类处理程序')
window.configure(bg='white')
# 创建大小，以及位移
formatForm(window, 500, 400)
# ------------------------------创建选择文件的容器------------------------------
frameSelectFile = Frame(window)
frameSelectFile.pack(padx=10, pady=10, fill=X)
frameSelectFile.configure(bg='white')
# 创建button
btn = Button(frameSelectFile, text='1.点击选择操作的文件', command=fileopen)
btn.pack(side=LEFT)
# 创建展示文件夹路径的输入框
filePath = StringVar()
filePathEntry = Entry(frameSelectFile, width=100, textvariable=filePath)
filePathEntry.pack(padx=8, side=RIGHT, fill=X)

# ------------------------------保存的路径------------------------------
frameSaveFile = Frame(window)
frameSaveFile.pack(padx=10, pady=10, fill=X)
frameSaveFile.configure(bg='white')
# 创建button
saveLabel = Label(frameSaveFile, text='2.保存的文件路径（请修改一下文件名，防止程序出错）')
saveLabel.pack(side=LEFT)
# 创建展示文件夹路径的输入框
filePathSave = StringVar()
saveEntry = Entry(frameSaveFile, width=100, textvariable=filePathSave)
saveEntry.pack(padx=8, side=RIGHT, fill=X)

runBtn = Button(window, text="运行", width='30', bg='red', command=run)
runBtn.pack(padx=10,pady=10, fill=X, side=BOTTOM)

# 窗体循环
window.mainloop()
