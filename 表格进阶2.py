import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook as open
from openpyxl.reader.excel import load_workbook
from excelbase import ExcelManager

# 加载表格
#
# wb = load_workbook('income.xlsx')
# print(wb.sheetnames)  # 输出所有工作表名称
# # 输出所有工作表
# for sheet in wb:
#     print(f"{sheet} + {sheet.title}")
# wb.create_sheet(title="按照名字创建sheet")  # 创建sheet表
# ws2 = wb.create_sheet(0)  # 创建工作表
#
# # 清除除了2018，2017，2016之外的sheet
# for sheet in wb.sheetnames:
#     if sheet.__eq__("2018") or sheet.__eq__("2017") or sheet.__eq__("2016"):
#         print("不清除" + sheet)
#     else:
#         wb.remove(wb[sheet])
#         print("清除" + sheet)
#
# # 读取工作表
# sheet2018 = wb['2018']
# print(f"总行数：{sheet2018.max_row}")
# print(f"总列数：{sheet2018.max_column}")
# # # 获取单元格a3
# # a3 = sheet2018["A3"]
# # # 获取单元格a3的内容
# # print(f"a3的数据是：{a3.value}")
#
# # 修改表格内容，以及输出表格内容
# sheet2018["A3"] = 123
# sheet2018["A3"] = 1234
# print(sheet2018["A3"].value)
#
# # 另外的方法输出内容,这里行和列不是从0开始的
# cell = sheet2018.cell(row=8, column=1)
# print(cell.value)
#
# # 访问多个单元格
# cell_range = sheet2018['A1':'A100']
# for cellRow in cell_range:
#     for cell in cellRow:
#         print(cell.value, end=",")
#
# # 整列修改
# # for i in range(1, 10):
# #     sheet2018['b' + i.__str__()] = "修改一下" + i.__str__()
#
# # 保存表格
# wb.save('income.xlsx')

# 加载表格
sheethjg = ExcelManager.getSheet(file='income.xlsx', sheetName="侯继国")

# # 获取行数
# print(excelManager.get_row_clo_num()[0])
# # 获取列数
# print(excelManager.get_row_clo_num()[1])
# # 获取第一行的数据
# first_row = excelManager.get_row_value(1)
# print(f"获取第一行的数据{first_row}")
# # 获取第二列的数据
# second_col = excelManager.get_col_value(2)
# print(f"获取第2列的数据{second_col}")
