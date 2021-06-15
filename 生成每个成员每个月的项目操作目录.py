from openpyxl import *
from openpyxl.cell import Cell
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.worksheet import Worksheet

# 不重复的姓名
nameSet: set = {""}
nameSet.remove("")

# 读取总表的数据
wbTotal: Workbook = load_workbook("/Users/jimmy/Desktop/工作簿.xlsx")

# ---------------------------读取全部数据---------------------------
wb1: Workbook = load_workbook("/Users/jimmy/Desktop/1月工时统计-final_modify.xlsx", data_only=True)
wb2: Workbook = load_workbook("/Users/jimmy/Desktop/2月工时统计-final_modify.xlsx", data_only=True)
wb3: Workbook = load_workbook("/Users/jimmy/Desktop/3月工时统计-final_modify.xlsx", data_only=True)
wb4: Workbook = load_workbook("/Users/jimmy/Desktop/4月工时统计-final_modify.xlsx", data_only=True)
wb5: Workbook = load_workbook("/Users/jimmy/Desktop/5月工时统计-final_modify.xlsx", data_only=True)
sheet1: Worksheet = wb1["数据源"]
sheet2: Worksheet = wb2["数据源"]
sheet3: Worksheet = wb3["数据源"]
sheet4: Worksheet = wb4["数据源"]
sheet5: Worksheet = wb5["数据源"]


# ---------------------------读取一月的数据---------------------------
# 根据sheet读取到相关的下标
def getIndexBysheet(sheet: Worksheet):
    index: dict = {}
    print("sheet的名称是：", sheet.title, "，共有列数是：", sheet.max_column, "，第一行最右边的单元格值是：",
          sheet.cell(row=1, column=sheet.max_column).value)
    for i in range(1, sheet.max_column + 1):
        cell: Cell = sheet.cell(row=1, column=i)
        cellValue: str = cell.value
        if cellValue is not None and "姓名".__contains__(cellValue):
            index.__setitem__("nameIndex", i)
        elif cellValue is not None and "项目".__contains__(cellValue):
            index.__setitem__("projectIndex", i)
        elif cellValue is not None and "实际工时数 (小时)".__contains__(cellValue):
            index.__setitem__("hourIndex", i)
    return index


index1 = getIndexBysheet(sheet1)
print("1月", index1)
# 通过循环得到姓名所在的列
nameIndex = index1.get("nameIndex")
projectIndex = index1.get("projectIndex")
hourIndex = index1.get("hourIndex")

index2 = getIndexBysheet(sheet2)
print("2月", index2)
index3 = getIndexBysheet(sheet3)
print("3月", index3)
index4 = getIndexBysheet(sheet4)
print("4月", index4)
index5 = getIndexBysheet(sheet5)
print("5月", index5)


# 取得所有sheet中的姓名一列的不重复数据
def addName(sheet: Worksheet, index: int):
    global nameSet
    for i in range(2, sheet.max_row + 1):
        cell: Cell = sheet.cell(row=i, column=index)
        name: str = cell.value
        nameSet.add(name)


print("正在整理第1个月的在职不重复人员")
addName(sheet1, index1.get("nameIndex"))
print("正在整理第2个月的在职不重复人员")
addName(sheet2, index2.get("nameIndex"))
print("正在整理第3个月的在职不重复人员")
addName(sheet3, index3.get("nameIndex"))
print("正在整理第4个月的在职不重复人员")
addName(sheet4, index4.get("nameIndex"))
print("正在整理第5个月的在职不重复人员")
addName(sheet5, index5.get("nameIndex"))
print("总的不重复人员是：", nameSet)
print("人员总数是：", len(nameSet))

print("处理表格内容中....")
# 根据唯一姓名，创建sheet
for name in nameSet:
    # 这两个是异常数据
    if name is None:
        print("发现异常数据None,跳过")
        continue
    if name.__eq__("#N/A"):
        print("发现异常数据#N/A,跳过")
        continue
    # 先创建
    wbTotal.create_sheet(title=name)
    # 获取到sheet
    nameSheet: Worksheet = wbTotal[name]

    # 创建固定数据
    # 设置高度
    nameSheet.row_dimensions[3].height = 30
    nameSheet.row_dimensions[4].height = 100
    nameSheet.row_dimensions[5].height = 100
    nameSheet.row_dimensions[6].height = 100
    nameSheet.row_dimensions[7].height = 100
    nameSheet.row_dimensions[8].height = 30

    # 设置宽度
    nameSheet.column_dimensions["b"].width = 20
    for i in range(1, 10):
        nameSheet.cell(column=2, row=i).alignment = Alignment(horizontal='center', vertical='center')
    nameSheet['b3'] = "战略项目"
    nameSheet['b3'].font = Font(bold=True)
    nameSheet['b4'] = "1"
    nameSheet['b5'] = "2"
    nameSheet['b6'] = "3"
    nameSheet['b7'] = "4"
    nameSheet['b8'] = "合计"

    nameSheet.column_dimensions["c"].width = 20
    nameSheet['c3'] = "工作内容和项目"
    nameSheet['c3'].alignment = Alignment(horizontal='center', vertical='center')
    nameSheet['c3'].font = Font(bold=True)

    nameSheet['c4'] = "平台优化改版"
    nameSheet['c5'] = "智能科创优化"
    nameSheet['c6'] = "TC_UK"
    nameSheet['c7'] = "酒店及旅行社相关"

    nameSheet.column_dimensions["d"].width = 40
    nameSheet['d3'] = "工作成果描述"
    nameSheet['d3'].font = Font(bold=True)
    nameSheet['d3'].alignment = Alignment(horizontal='center', vertical='center')

    nameSheet.column_dimensions["e"].width = 40
    nameSheet['e3'] = "对应交付"
    nameSheet['e3'].font = Font(bold=True)
    nameSheet['e3'].alignment = Alignment(horizontal='center', vertical='center')

    nameSheet.column_dimensions["f"].width = 10
    nameSheet['f3'] = "1月"
    nameSheet['f3'].font = Font(bold=True)
    nameSheet['f3'].alignment = Alignment(horizontal='center', vertical='center')

    # 需要在1月的内容里面填数据，首先是姓名 根据是c4,c5,c6,c7,填写f4,f5,f6,f7
    sheet1_rows_value = sheet1.rows
    for row in sheet1_rows_value:
        sheet1_name: str = row[nameIndex - 1].value
        # 判断两个名字相同，说明1月份的当前姓名的行已经找到了。
        if sheet1_name.__eq__(name):
            # 接下来就是要找对应的项目，将项目的实际工时累加起来，填入到excel中
            # 取出姓名一致时的项目
            projecName: str = row[projectIndex - 1].value
            # print(projecName)
            # print(nameSheet['c4'].value)
            if projecName.__eq__(nameSheet['c4'].value):  # 平台改版优化
                originHour: int = row[hourIndex - 1].value
                # 如果数据是空的
                if None is nameSheet['f4'].value:
                    nameSheet['f4'] = 0
                # 进行当前人，当前项目的累加操作
                nameSheet['f4'] = nameSheet['f4'].value + originHour
            elif projecName.__eq__(nameSheet['c5'].value):  # 智能科创优化
                originHour: int = row[hourIndex - 1].value
                # 如果数据是空的
                if None is nameSheet['f5'].value:
                    nameSheet['f5'] = 0
                # 进行当前人，当前项目的累加操作
                nameSheet['f5'] = nameSheet['f5'].value + originHour
            elif projecName.__eq__(nameSheet['c6'].value):  # TC_UK
                originHour: int = row[hourIndex - 1].value
                # 如果数据是空的
                if None is nameSheet['f6'].value:
                    nameSheet['f6'] = 0
                # 进行当前人，当前项目的累加操作
                nameSheet['f6'] = nameSheet['f6'].value + originHour
            elif projecName.__eq__(nameSheet['c7'].value):  # 酒店及旅行社相关
                originHour: int = row[hourIndex - 1].value
                # 如果数据是空的
                if None is nameSheet['f7'].value:
                    nameSheet['f7'] = 0
                # 进行当前人，当前项目的累加操作
                nameSheet['f7'] = nameSheet['f7'].value + originHour
    # 填写f8 1月的总计
    oneMonthTotal: int = 0
    for i in range(4, 8):
        value = nameSheet.cell(row=i, column=6).value
        if value is None:
            continue
        oneMonthTotal += value

    # print(name, oneMonthTotal)
    if oneMonthTotal != 0:
        nameSheet['f8'] = oneMonthTotal

    nameSheet.column_dimensions["g"].width = 10
    nameSheet['g3'] = "2月"
    nameSheet['g3'].font = Font(bold=True)
    nameSheet['g3'].alignment = Alignment(horizontal='center', vertical='center')

    nameSheet.column_dimensions["h"].width = 10
    nameSheet['h3'] = "3月"
    nameSheet['h3'].font = Font(bold=True)
    nameSheet['h3'].alignment = Alignment(horizontal='center', vertical='center')

    nameSheet.column_dimensions["i"].width = 10
    nameSheet['i3'] = "4月"
    nameSheet['i3'].font = Font(bold=True)
    nameSheet['i3'].alignment = Alignment(horizontal='center', vertical='center')

    nameSheet.column_dimensions["j"].width = 10
    nameSheet['j3'] = "5月"
    nameSheet['j3'].font = Font(bold=True)
    nameSheet['j3'].alignment = Alignment(horizontal='center', vertical='center')

wbTotal.save("/Users/jimmy/Desktop/2021上半年.xlsx")
